import openpyxl as px
import language.language_converter as lc
class XlsxCyrilicLatinConverter:
    EXTENSION = "LATIN"
    def __init__(self, work_book_name):
        self.work_book_name = work_book_name

    def convert_cyrilic_to_latin(self):
        work_book = px.load_workbook(filename=self.work_book_name)
        sheet_names = work_book.get_sheet_names();
        for sheet_name in sheet_names:
            sheet = work_book[sheet_name]
            for row_idx in range(1, sheet.max_row):
                for col_idx in range(1, sheet.max_column):
                    cell_value = sheet.cell(column = col_idx, row=row_idx).value
                    latin_cell_value = lc.CyrilicToLatin.convertCyrilicToLatin(cell_value)
                    if (type(latin_cell_value) is str):
                        sheet.cell(column = col_idx, row=row_idx).set_explicit_value(latin_cell_value)

        work_book.save("LATIN" + self.work_book_name)
        work_book.close()