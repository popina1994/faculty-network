import openpyxl as px
import re
from openpyxl import Workbook
from graph.author import Author
from graph.work import Work

class GraphXlsxConverter:
    SOURCE_COLUMN_IDX = 1
    TARGET_COLUMN_IDX = 2
    COLUMN_NAMES_ROW_ID = 1

    COLUMN_NAME_AUTHORS_FULL_NAME = "konkatenirano"
    COLUMN_NAME_OF_AUTHORS_OPT_1 = "autori"
    COLUMN_NAME_OF_AUTHORS_OPT_2 = "autori2"

    COLUMN_NAME_OF_YEAR = "godina"
    COLUMN_NAME_CONF_OR_MAG_OPT_1 = "cas_naslov"
    COLUMN_NAME_CONF_OR_MAG_OPT_2 = "naziv_skupa"
    COLUMN_NAME_DEPARTMENT = "organizaciona_jedinica_naziv"

    COLUMN_NAME_SOURCE_NODE = "Source"
    COLUMN_NAME_DEST_NODE = "Target"

    FILE_NAME_WORK_PER_AUTHOR = "Broj radova po naucniku.txt"
    FILE_NAME_WORK_PER_WORK = "Radovi.txt"
    FILE_NAME_PRODUCTIVITY = "oblast-godina.txt"

    authors = {}
    works = {}

    def find_authors_full_name_column_idx(sheet):
        if (sheet.cell(column = 1, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID) is None):
            return 0
        for col_idx in range(1, sheet.max_column):
            cell = sheet.cell(column = col_idx, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID)
            if (cell.value == GraphXlsxConverter.COLUMN_NAME_AUTHORS_FULL_NAME):
                return col_idx
        return 0

    def find_department_column_idx(sheet):
        if (sheet.cell(column = 1, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID) is None):
            return 0
        for col_idx in range(1, sheet.max_column):
            cell = sheet.cell(column = col_idx, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID)
            if (cell.value == GraphXlsxConverter.COLUMN_NAME_DEPARTMENT):
                return col_idx
        return 0

    def init(work_book_authors_name):
        work_book_read = px.load_workbook(filename=work_book_authors_name)
        sheet_names = work_book_read.get_sheet_names();
        for sheet_name in sheet_names:
            #Initialization of idxs
            sheet = work_book_read[sheet_name]
            authors_full_name_col_idx = GraphXlsxConverter.find_authors_full_name_column_idx(sheet)
            if (authors_full_name_col_idx == 0):
                continue
            authors_department_name_col_idx = GraphXlsxConverter.find_department_column_idx(sheet)

            for row_idx in range(GraphXlsxConverter.COLUMN_NAMES_ROW_ID + 1, sheet.max_row):
                cell_value = sheet.cell(column = authors_full_name_col_idx, row = row_idx).value
                if ((cell_value == "") or (cell_value is None)):
                    continue
                author = cell_value.strip()
                department = sheet.cell(column = authors_department_name_col_idx, row = row_idx).value.strip()
                GraphXlsxConverter.authors[author] = Author(author, department)

    def __init__(self, work_book_read_name):
        self.work_book_read_name = work_book_read_name

    def find_autors_column_idx(sheet):
        if (sheet.cell(column = 1, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID) is None):
            return "", 0
        for col_idx in range(1, sheet.max_column):
            cell = sheet.cell(column = col_idx, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID)
            if (cell.value == GraphXlsxConverter.COLUMN_NAME_OF_AUTHORS_OPT_1):
                return GraphXlsxConverter.COLUMN_NAME_OF_AUTHORS_OPT_1, col_idx
            if (cell.value == GraphXlsxConverter.COLUMN_NAME_OF_AUTHORS_OPT_2):
                return GraphXlsxConverter.COLUMN_NAME_OF_AUTHORS_OPT_2, col_idx
        return "", 0

    def find_year_column_idx(sheet):
        if (sheet.cell(column = 1, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID) is None):
            return "", 0
        for col_idx in range(1, sheet.max_column):
            cell = sheet.cell(column = col_idx, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID)
            if (cell.value == GraphXlsxConverter.COLUMN_NAME_OF_YEAR):
                return GraphXlsxConverter.COLUMN_NAME_OF_YEAR, col_idx

        return "", 0

    def find_work_column_idx(sheet):
        if (sheet.cell(column = 1, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID) is None):
            return "", 0
        for col_idx in range(1, sheet.max_column):
            cell = sheet.cell(column = col_idx, row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID)
            if (cell.value == GraphXlsxConverter.COLUMN_NAME_CONF_OR_MAG_OPT_1):
                return False, col_idx
            if (cell.value == GraphXlsxConverter.COLUMN_NAME_CONF_OR_MAG_OPT_2):
                return True, col_idx
        return "", 0

    def init_sheet_names(sheet):
        sheet.cell(row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID, 
                   column = GraphXlsxConverter.SOURCE_COLUMN_IDX).value = GraphXlsxConverter.COLUMN_NAME_SOURCE_NODE
        sheet.cell(row = GraphXlsxConverter.COLUMN_NAMES_ROW_ID, 
                   column = GraphXlsxConverter.TARGET_COLUMN_IDX).value = GraphXlsxConverter.COLUMN_NAME_DEST_NODE

    def generate_edges_of_authors(self, sheet, authors, conference):
        for author_1 in authors:
            authorClass = GraphXlsxConverter.authors[author_1]
            authorClass.work_num = authorClass.work_num + 1
            authorClass.conf_num = authorClass.conf_num + int(conference)
            for author_2 in authors:
                sheet.cell(row = self.row_write_idx, column = GraphXlsxConverter.SOURCE_COLUMN_IDX).value = author_1
                sheet.cell(row = self.row_write_idx, column = GraphXlsxConverter.TARGET_COLUMN_IDX).value = author_2
                self.row_write_idx += 1

    def add_work(work_name, isConference, authors, year):
        if (type(work_name) is type(None)):
            work_name = "Nema"
        work_name = work_name.replace(":", "")
        work_name_u = ("konf_" if isConference else "mag_") + work_name
        if (work_name_u not in GraphXlsxConverter.works):
            work = Work(work_name_u, isConference) 
            GraphXlsxConverter.works[work_name_u] = work
        GraphXlsxConverter.works[work_name_u].add_issue(authors, year)

    def read_year_from_sheet(sheet, row_idx, year_col_idx):
        year_value = sheet.cell(column = year_col_idx, row = row_idx).value
        if (type(year_value) is int):
            year_value = int(year_value)
        else:
            """
            if (type(year_value) is type(None)):
                print("None")
            else:
                print("Irregular" + year_value + "\n")
            """
            return False, -1 
        if ((year_value < 2000) or (year_value > 2016)):
            return False, -1
        return True, year_value

    def convert_xlsx_to_graph(self):
        work_book_read = px.load_workbook(filename=self.work_book_read_name)
        work_book_write = Workbook()
        sheet_write_active = work_book_write.active
        sheet_names = work_book_read.get_sheet_names();
        for sheet_name in sheet_names:
            sheet = work_book_read[sheet_name]
            option, authors_col_idx = GraphXlsxConverter.find_autors_column_idx(sheet)
            _, year_col_idx = GraphXlsxConverter.find_year_column_idx(sheet)
            isConference, work_col_idx = GraphXlsxConverter.find_work_column_idx(sheet)
            if (authors_col_idx == 0):
                continue
            
            GraphXlsxConverter.init_sheet_names(sheet_write_active)
            self.row_write_idx = GraphXlsxConverter.COLUMN_NAMES_ROW_ID + 1
            for row_idx in range(GraphXlsxConverter.COLUMN_NAMES_ROW_ID + 1, sheet.max_row):
                isOk, year = GraphXlsxConverter.read_year_from_sheet(sheet, row_idx, year_col_idx)
                if not isOk: continue

                cell_value = sheet.cell(column = authors_col_idx, row = row_idx).value
                if ((cell_value == "") or (cell_value is None)):
                    continue
                
                if (option == GraphXlsxConverter.COLUMN_NAME_OF_AUTHORS_OPT_2):
                    # Remove the first {, and at the end }"
                    cell_value =  "\"" + cell_value[1:-2];

                list_of_authors_not_formated = cell_value.split(",")
                list_of_authors = []
                for author_not_format in list_of_authors_not_formated:
                    author = author_not_format.strip()
                    author = re.sub(' +',' ',author);
                    
                    if (option == GraphXlsxConverter.COLUMN_NAME_OF_AUTHORS_OPT_2):
                        # remove ""
                        author = author[1 : -1]
                    if (author.find(".") != -1):
                        continue
                    if (author not in GraphXlsxConverter.authors):
                        continue
                    list_of_authors.append(author)
                #there is no list
                if not list_of_authors:
                    continue
                self.generate_edges_of_authors(sheet_write_active, list_of_authors, isConference)
                work_name = sheet.cell(column = work_col_idx, row = row_idx).value
                GraphXlsxConverter.add_work(work_name, isConference, list_of_authors, year)

        work_book_write.save("GRAPH_" + self.work_book_read_name)
        work_book_read.close()


    def sort_authors_by_work_number():
        list_authors = list(GraphXlsxConverter.authors.values())
        list_authors.sort()
        list_authors.reverse()
        with open(GraphXlsxConverter.FILE_NAME_WORK_PER_AUTHOR, 'w', encoding="utf-8") as f:
            f.write('Ime i prezime, Katedra, Ukupno, Konferencije, Magazini\n')
            for author in list_authors:
                f.write('{0!s}\n'.format(author))
        return list_authors

    def sort_work_by_name_year():
        list_work = list(GraphXlsxConverter.works.values())
        list_work.sort()

        with open(GraphXlsxConverter.FILE_NAME_WORK_PER_WORK, 'w', encoding="utf-8") as f:
            f.write('Ime: Autora: Godina\n')
            for work in list_work:
                work.sort_issues()
                for issue in work.issues:
                    f.write('{0!s}: {1!s}\n'.format(work.name, issue))
        return list_work

    def output_department_year_fraction():
        list_work = list(GraphXlsxConverter.works.values())

        with open(GraphXlsxConverter.FILE_NAME_PRODUCTIVITY, 'w', encoding="utf-8") as f:
            f.write('Departman: Godina: Frakcija\n');
            for work in list_work:
                for issue in work.issues:
                    frac = 1 / len(issue.authors)
                    for author in issue.authors:
                        authorClass = GraphXlsxConverter.authors[author]
                        f.write('{0!s}: {1!s}: {2!s}\n'.format(authorClass.department, 
                                                       issue.year, frac))
        return list_work


            

        